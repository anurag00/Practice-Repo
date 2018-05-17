import openpyxl, logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
logging.debug(wb.worksheets)
wb.create_sheet(title='Tables',index=0)
logging.debug(wb.worksheets)
sheet = wb.worksheets[0]

#First row Frozen and BOLD
fontObj = Font(name='Times New Roman', bold=True)
for colNum in range(1,10):
    sheet.column_dimensions[get_column_letter(colNum)].width = 15
    sheet.cell(row=1, column=colNum).font = fontObj
    sheet.cell(row=1,column=colNum).value = "Number = " + str(colNum)
sheet = wb.active
sheet.freeze_panes = 'a2'

#Rest of the rows (tables upto 20)
for colNum in range(1,sheet.max_column+1):
    for rowNum in range(2,22):
        sheet.cell(row = rowNum, column = colNum).value = (rowNum-1)*colNum

#merge last row
sheet.merge_cells('a23:i23')

#SUM function for all the tables
for colNum in range(1,10):
    colAlpha = get_column_letter(colNum)
    formulaEx = '=SUM('+colAlpha+'2:'+colAlpha+'21)'
    #=SUM(A2:A21)
    logging.debug(formulaEx)
    sheet.cell(row=22,column=colNum).value = formulaEx

fontObj = Font(name='Ariel', bold=True)
sheet['a23'].font = fontObj
sheet['a23'].value = "Charts Bitches"

#Creating Charts
refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=2, max_col=1, max_row=21)
seriesObj = openpyxl.chart.Series(refObj, title='First series')
chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)
sheet.add_chart(chartObj, 'a24')
logging.debug("Ended chart")
logging.debug(sheet.max_row)
logging.debug(sheet.max_column)

wb.save(r"G:\Pycode\Automate\everythingExcel.xlsx")
