import openpyxl
from openpyxl.utils import get_column_letter
workbook = openpyxl.load_workbook("G:\Pycode\Automate\example.xlsx")
print(type(workbook))
#sheet = workbook.get_sheet_by_name("Sheet1")
workbook.get_sheet_names()[0]
print("Number of worksheets = ",len(workbook.worksheets))
print(list(workbook.worksheets))
sheet = workbook.worksheets[0]
for x in range(1,sheet.max_row+1):
    for y in range(1,sheet.max_column+1):
        print(sheet.cell(row = x, column = y).value,' ',end='')  #sheet.cell(row = x, column = y).coordinate
    print('')
print(get_column_letter(775))
print(list(sheet['a1':'b7']))

for rowOfCells in sheet['a1':'b7']:
    for cellObj in rowOfCells:
        print(cellObj.coordinate,cellObj.value,end=' ')
    print('')